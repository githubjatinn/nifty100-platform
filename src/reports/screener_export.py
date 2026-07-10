"""Generates output/screener_output.xlsx: one sheet per preset, sorted by composite score, threshold-coloured cells."""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

OUTPUT_PATH = "output/screener_output.xlsx"

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FONT = Font(bold=True)

DISPLAY_METRIC_COLUMNS = [
    "return_on_equity_pct", "return_on_capital_pct", "net_profit_margin_pct",
    "operating_profit_margin_pct", "ebit_margin_pct", "return_on_assets_pct",
    "debt_to_equity", "interest_coverage", "net_debt_to_ebitda",
    "free_cash_flow_cr", "fcf_conversion_pct", "cash_from_operations_cr",
    "revenue_cagr_5yr", "pat_cagr_5yr", "eps_cagr_5yr",
    "dividend_payout_ratio_pct", "asset_turnover", "fixed_asset_turnover",
    "pe_ratio", "pb_ratio", "composite_quality_score",
]

FILTER_KEY_TO_COLUMN = {
    "roe_min": ("return_on_equity_pct", "min"),
    "de_max": ("debt_to_equity", "max"),
    "fcf_min": ("free_cash_flow_cr", "min"),
    "revenue_cagr_5yr_min": ("revenue_cagr_5yr", "min"),
    "pat_cagr_5yr_min": ("pat_cagr_5yr", "min"),
    "opm_min": ("operating_profit_margin_pct", "min"),
    "pe_max": ("pe_ratio", "max"),
    "pb_max": ("pb_ratio", "max"),
    "dividend_yield_min": ("dividend_yield_pct", "min"),
    "icr_min": ("interest_coverage", "min"),
    "market_cap_min": ("market_cap_crore", "min"),
    "dividend_payout_min": ("dividend_payout_ratio_pct", "min"),
    "dividend_payout_max": ("dividend_payout_ratio_pct", "max"),
    "revenue_min_cr": ("market_cap_crore", "min"),
}


def _preset_thresholds(preset_dict):
    """Converts a preset's raw config dict into {column: (direction, threshold)} pairs, skipping non-numeric flags."""
    thresholds = {}
    for key, value in preset_dict.items():
        if key not in FILTER_KEY_TO_COLUMN or not isinstance(value, (int, float)):
            continue
        column, direction = FILTER_KEY_TO_COLUMN[key]
        thresholds[column] = (direction, value)
    return thresholds


def _cell_fill(column, value, thresholds):
    """Returns green if the cell meets its preset's threshold for that column, red if it fails, None if not applicable."""
    if column not in thresholds or value in (None, "") or pd.isna(value):
        return None
    direction, threshold = thresholds[column]
    passes = value >= threshold if direction == "min" else value <= threshold
    return GREEN_FILL if passes else RED_FILL


def _write_preset_sheet(wb, preset_name, preset_df, preset_dict, metric_columns=DISPLAY_METRIC_COLUMNS):
    """Writes one preset's companies (sorted by composite score desc) with threshold-based green/red cell coding."""
    ws = wb.create_sheet(title=preset_name[:31])
    thresholds = _preset_thresholds(preset_dict)

    display_columns = ["company_id", "company_name"] + [c for c in metric_columns if c in preset_df.columns]
    ws.append(display_columns)
    for cell in ws[1]:
        cell.font = HEADER_FONT

    sorted_df = preset_df.sort_values("composite_quality_score", ascending=False) \
        if "composite_quality_score" in preset_df.columns else preset_df

    for _, row in sorted_df.iterrows():
        ws.append([row.get(c, "") for c in display_columns])
        row_idx = ws.max_row
        for col_idx, metric in enumerate(display_columns[2:], start=3):
            fill = _cell_fill(metric, row.get(metric), thresholds)
            if fill:
                ws.cell(row=row_idx, column=col_idx).fill = fill

    if len(sorted_df) == 0:
        ws.append(["No companies matched this preset."])

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)


def build_screener_output_workbook(preset_results, config, output_path=OUTPUT_PATH):
    """Builds screener_output.xlsx: one sheet per preset, sorted by composite score, threshold-coloured cells."""
    preset_key_map = {
        "Quality Compounder": "quality_compounder",
        "Value Pick": "value_pick",
        "Growth Accelerator": "growth_accelerator",
        "Dividend Champion": "dividend_champion",
        "Debt-Free Blue Chip": "debt_free_blue_chip",
        "Turnaround Watch": "turnaround_watch",
    }

    wb = Workbook()
    wb.remove(wb.active)

    for preset_name, preset_df in preset_results.items():
        preset_key = preset_key_map.get(preset_name)
        preset_dict = config["presets"].get(preset_key, {})
        _write_preset_sheet(wb, preset_name, preset_df, preset_dict)

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path
"""Generates output/peer_comparison.xlsx: one colour-coded sheet per peer group, metrics + percentile ranks."""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

OUTPUT_PATH = "output/peer_comparison.xlsx"

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
HEADER_FONT = Font(bold=True)

DEFAULT_METRIC_COLUMNS = [
    "return_on_equity_pct", "return_on_capital_pct", "net_profit_margin_pct",
    "operating_profit_margin_pct", "ebit_margin_pct", "return_on_assets_pct",
    "debt_to_equity", "interest_coverage", "net_debt_to_ebitda",
    "free_cash_flow_cr", "fcf_conversion_pct", "cash_from_operations_cr",
    "revenue_cagr_5yr", "pat_cagr_5yr", "eps_cagr_5yr",
    "dividend_payout_ratio_pct", "asset_turnover", "fixed_asset_turnover",
    "pe_ratio", "pb_ratio",
]


def _percentile_fill(pct):
    """Returns the cell fill colour for a percentile rank: green >=75, yellow 25-75, red <25."""
    if pct is None:
        return None
    if pct >= 75:
        return GREEN_FILL
    if pct >= 25:
        return YELLOW_FILL
    return RED_FILL


def _write_sheet(wb, sheet_name, group_df, metric_columns, percentile_lookup):
    """Writes one peer group's comparison table with percentile colour-coding and a median summary row."""
    ws = wb.create_sheet(title=str(sheet_name)[:31])

    display_columns = ["company_id", "company_name"] + metric_columns
    ws.append(display_columns)
    for cell in ws[1]:
        cell.font = HEADER_FONT

    for _, row in group_df.iterrows():
        ws.append([row.get(c, "") for c in display_columns])
        row_idx = ws.max_row
        if row.get("is_benchmark") in (1, True):
            for cell in ws[row_idx]:
                cell.fill = GOLD_FILL
        for col_idx, metric in enumerate(metric_columns, start=3):
            pct = percentile_lookup.get((row["company_id"], metric))
            fill = _percentile_fill(pct)
            if fill:
                ws.cell(row=row_idx, column=col_idx).fill = fill

    median_row = ["", "Peer Group Median"] + [
        pd.to_numeric(group_df[m], errors="coerce").median() if m in group_df.columns else "" for m in metric_columns
    ]
    ws.append(median_row)
    for cell in ws[ws.max_row]:
        cell.font = HEADER_FONT

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)


def build_peer_comparison_workbook(df, percentiles_df, metric_columns=None, output_path=OUTPUT_PATH):
    """Builds the multi-sheet workbook: one sheet per peer group, using peer_percentiles data for colour-coding.

    Benchmark rows are highlighted using the is_benchmark flag already present in the peer_groups table.
    """
    metric_columns = metric_columns or [c for c in DEFAULT_METRIC_COLUMNS if c in df.columns]
    percentile_lookup = {
        (row["company_id"], row["metric"]): row["percentile_rank"]
        for _, row in percentiles_df.iterrows()
        if row["percentile_rank"] is not None
    }

    wb = Workbook()
    wb.remove(wb.active)

    for group_name, group_df in df.groupby("peer_group_name", dropna=False):
        if pd.isna(group_name):
            continue  # skip companies with no peer group assigned — not part of the 11 defined groups
        _write_sheet(wb, group_name, group_df, metric_columns, percentile_lookup)

    import os
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return output_path